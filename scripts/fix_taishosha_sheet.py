# -*- coding: utf-8 -*-
"""「対象者」シートの #REF! を修復する。

旧数式は削除済みの「移管一覧」シート(C列=オーナーのメールアドレス、
E列=完了ステータス)を参照しており復元不能。代わりに現存する
ユニット4シートから再定義した値を静的に書き込む:

- 対応数 = ファイルオーナー一覧(col E)に本人メールが登場する案件数
- 完了数 = うち Migration Status(col M)が COMPLETED の案件数
- 完了率 = 完了数/対応数 (対応数0なら空欄)

実行前に <元ファイル名>_backup.xlsx を自動作成する。

Usage:
    python scripts/fix_taishosha_sheet.py [xlsx_path]
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = REPO_ROOT / "ドライブ移管対応_進捗管理20260706.xlsx"
UNIT_SHEETS = ["第1ユニット", "第2ユニット", "第3ユニット", "アソシエイト"]

COL_OWNERS = 4   # ファイルオーナー一覧 (0-based)
COL_STATUS = 12  # Migration Status (0-based)


def collect_owner_stats(wb) -> dict[str, list[int]]:
    """email -> [対応数, 完了数]"""
    stats: dict[str, list[int]] = {}
    for sheet in UNIT_SHEETS:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None:
                continue
            owners_raw = row[COL_OWNERS] if len(row) > COL_OWNERS else None
            if not owners_raw:
                continue
            status = row[COL_STATUS] if len(row) > COL_STATUS else None
            done = status == "COMPLETED"
            for email in str(owners_raw).splitlines():
                email = email.strip().lower()
                if not email or "@" not in email:
                    continue
                entry = stats.setdefault(email, [0, 0])
                entry[0] += 1
                if done:
                    entry[1] += 1
    return stats


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        sys.exit(f"xlsx が見つかりません: {xlsx_path}")

    backup = xlsx_path.with_name(xlsx_path.stem + "_backup.xlsx")
    if not backup.exists():
        shutil.copy2(xlsx_path, backup)
        print(f"バックアップ作成: {backup.name}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    stats = collect_owner_stats(wb)

    ws = wb["対象者"]
    updated = 0
    nonzero = 0
    for row in ws.iter_rows(min_row=3):
        email_cell = row[1]  # B列
        if not email_cell.value:
            continue
        email = str(email_cell.value).strip().lower()
        taio, kanryo = stats.get(email, [0, 0])
        row[3].value = taio                      # D: 対応数
        row[4].value = kanryo                    # E: 完了数
        row[5].value = (kanryo / taio) if taio else None  # F: 完了率
        row[5].number_format = "0%"
        updated += 1
        if taio:
            nonzero += 1

    ws.cell(row=1, column=2).value = (
        "対応数/完了数は 2026-07-07 に再定義: "
        "ユニット4シートのファイルオーナー一覧に登場する案件数 / うち Migration Status=COMPLETED"
        "(旧「移管一覧」シート削除により #REF! となっていたため)"
    )

    wb.save(xlsx_path)
    print(f"対象者 {updated} 行を更新(対応数1件以上: {nonzero} 人)")
    top = sorted(stats.items(), key=lambda x: -x[1][0])[:10]
    print("対応数 上位10:")
    for email, (t, k) in top:
        print(f"  {email}: 対応 {t} / 完了 {k}")


if __name__ == "__main__":
    main()
